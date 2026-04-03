"""
stage1_app.py
=============
Input:  331 PDF + TXT.BKMVDATA (ZIP or folder) + header code
Output: Excel — Account Number / Account Name / Opening Balance / Closing Balance

Exact extraction logic (confirmed from file inspection):

B100 records — account master:
  account_number = record[172:178].strip()
  (name and opening balance: inspect positions and print what is used)

C100 records — movements:
  matches = re.findall(r'[+-]\\d+', record)
  account_number = abs(int(matches[-1]))          ← last signed field
  amount         = abs(int(matches[-2]))          ← second-to-last signed field
  sign           = +1 if matches[-1] starts with '+' else -1   (debit / credit)
  net_movement   = sign × amount

Closing Balance = Opening Balance + Σ net_movements for that account (year 2025)

Sign convention: positive = debit, negative = credit
"""

from __future__ import annotations

import io
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import streamlit as st

_HERE = Path(__file__).parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

import openpyxl
from openpyxl.styles import Font, PatternFill
from parsers.report331_parser import parse_331_pdf

TARGET_YEAR = 2025


# ── helpers ───────────────────────────────────────────────────────────────────

def _norm(key) -> str:
    """Normalize account key: strip, remove leading zeros."""
    return str(key).strip().lstrip("0") or "0"


def _parse_amount(raw: str) -> float:
    raw = raw.strip().replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _find_bkmvdata(root: str) -> str | None:
    NAMES = {"TXT.BKMVDATA", "BKMVDATA.TXT", "BKMVDATA"}
    for dirpath, _, files in os.walk(root):
        for fname in files:
            if fname.upper() in NAMES:
                return os.path.join(dirpath, fname)
    return None


# ── page layout ───────────────────────────────────────────────────────────────

st.set_page_config(page_title="Ageing – Balances", layout="wide")
st.title("Account Balances from TXT.BKMVDATA")
st.divider()

col_pdf, col_bkmv = st.columns(2)

with col_pdf:
    st.subheader("1. 331 PDF")
    pdf_file = st.file_uploader("Upload 331 PDF", type=["pdf"], key="pdf_up")
    if pdf_file:
        st.success(f"{pdf_file.name} ({pdf_file.size:,} bytes)")

with col_bkmv:
    st.subheader("2. Uniform Structure")
    method = st.radio("Input:", ["ZIP file", "Folder path"], horizontal=True)
    zip_bytes   = None
    folder_path = ""
    if method == "ZIP file":
        zf = st.file_uploader("Upload ZIP", type=["zip"], key="zip_up")
        if zf:
            zip_bytes = zf.read()
            st.success(f"{zf.name} ({len(zip_bytes):,} bytes)")
    else:
        folder_path = st.text_input("Folder path", key="folder_in")
        if folder_path and os.path.isdir(folder_path):
            st.success("Folder found")

st.divider()
header_code = st.text_input("Header Code (e.g. 1342)", placeholder="1342", max_chars=20)

bkmv_ready = bool(zip_bytes) or bool(folder_path.strip())
confirm = st.button(
    "✅ Confirm", type="primary",
    disabled=(pdf_file is None or not bkmv_ready or not header_code.strip()),
)

# ── processing ────────────────────────────────────────────────────────────────

if confirm:
    hcode   = header_code.strip()
    tmp_dir = None

    pdf_data = pdf_file.read()
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False, prefix="331_") as tmp:
        tmp.write(pdf_data)
        pdf_path = tmp.name

    try:
        progress = st.progress(0, text="Starting…")

        # ── Step 1: PDF → account numbers ────────────────────────────────
        progress.progress(10, text="Parsing 331 PDF…")
        report = parse_331_pdf(pdf_path, hcode)

        if not report.header_line_raw:
            st.error(f"Header '{hcode}' not found in PDF.")
            st.stop()

        pdf_accounts = [acc.account_number for acc in report.accounts]
        pdf_norms    = {_norm(a) for a in pdf_accounts}

        st.success(
            f"331 PDF – **{hcode} – {report.header_name}** – "
            f"**{len(pdf_accounts)} account numbers**"
        )

        # ── Step 2: locate TXT.BKMVDATA ──────────────────────────────────
        progress.progress(20, text="Locating TXT.BKMVDATA…")
        if zip_bytes:
            tmp_dir = tempfile.mkdtemp(prefix="bkmv_")
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                zf.extractall(tmp_dir)
            root = tmp_dir
        else:
            root = folder_path.strip()

        bkmvdata_path = _find_bkmvdata(root)
        if not bkmvdata_path:
            st.error("TXT.BKMVDATA not found.")
            st.stop()
        st.info(f"File: `{bkmvdata_path}`")

        # ── Step 3: single-pass parse of B100 + C100 ─────────────────────
        progress.progress(35, text="Parsing TXT.BKMVDATA…")

        # accounts_master[norm_key] = {number, name, opening}
        accounts_master: dict[str, dict] = {}
        # movements_net[norm_key] = signed sum for TARGET_YEAR
        movements_net: dict[str, float] = {}

        b100_count = c100_count = b100_field_samples = 0
        b100_samples = []   # raw bytes of first 3 B100 lines for debug
        c100_samples = []   # signed-field lists of first 3 C100 lines for debug

        with open(bkmvdata_path, "rb") as f:
            for raw_line in f:
                if not raw_line.strip():
                    continue

                prefix = raw_line[:20]

                # ── B100: account master ──────────────────────────────
                if b"B100" in prefix:
                    b100_count += 1
                    decoded = raw_line.decode("latin-1", errors="replace")

                    # Account number: confirmed position 172:178
                    acc_num = decoded[172:178].strip()

                    # Capture first 3 records for field-position debug
                    if b100_count <= 3:
                        b100_samples.append(decoded)

                    if acc_num:
                        accounts_master[_norm(acc_num)] = {
                            "number":  acc_num,
                            "name":    "",       # position TBD — shown in debug
                            "opening": 0.0,      # position TBD — shown in debug
                            "_raw":    decoded,
                        }

                # ── C100: movements ───────────────────────────────────
                elif b"C100" in prefix:
                    c100_count += 1
                    decoded = raw_line.decode("latin-1", errors="replace")

                    # Date field: try positions 8-16 (YYYYMMDD)
                    # (show in debug if wrong)
                    date_raw = decoded[8:16].strip()
                    if not date_raw.startswith(str(TARGET_YEAR)):
                        continue

                    matches = re.findall(r'[+-]\d+', decoded)

                    if len(matches) < 2:
                        continue   # need at least amount + account

                    # Last signed field = account number
                    acc_signed  = int(matches[-1])
                    acc_num     = abs(acc_signed)
                    acc_norm    = _norm(acc_num)

                    # Second-to-last signed field = movement amount
                    amt_signed  = int(matches[-2])
                    amount      = abs(amt_signed)

                    # Sign of account field: + = debit, - = credit
                    net = +amount if acc_signed > 0 else -amount

                    movements_net[acc_norm] = movements_net.get(acc_norm, 0.0) + net

                    if c100_count <= 3:
                        c100_samples.append((decoded.strip(), matches))

        st.success(
            f"TXT.BKMVDATA – **{b100_count:,} B100** records  |  "
            f"**{c100_count:,} C100** records  |  "
            f"{len(movements_net):,} accounts with {TARGET_YEAR} movements"
        )

        # ── B100 field-position debug ─────────────────────────────────────
        with st.expander("B100 raw field layout (first 3 records — verify name + opening balance positions)"):
            for i, rec in enumerate(b100_samples, 1):
                st.text(f"Record {i}:")
                st.text(f"  [172:178] acc_num  = {rec[172:178]!r}")
                # Show nearby slices to find name and opening balance
                for start in range(0, min(len(rec), 200), 10):
                    chunk = rec[start:start+10]
                    if chunk.strip():
                        st.text(f"  [{start:3d}:{start+10:3d}] = {chunk!r}")
                st.text("---")

        # ── C100 field debug ──────────────────────────────────────────────
        with st.expander("C100 signed fields (first 3 matching records)"):
            for decoded, matches in c100_samples:
                st.text(f"LINE: {decoded[:120]!r}")
                st.text(f"  signed fields: {matches}")
                if len(matches) >= 2:
                    st.text(f"  → account  = {abs(int(matches[-1]))} (last)")
                    st.text(f"  → amount   = {abs(int(matches[-2]))} (second-to-last)")
                    st.text(f"  → debit/credit = {'DEBIT (+)' if int(matches[-1]) > 0 else 'CREDIT (-)'}")
                st.text("---")

        # ── Step 4: build results ─────────────────────────────────────────
        progress.progress(70, text="Calculating balances…")

        results   = []
        not_found = []

        for acc_num in pdf_accounts:
            key  = _norm(acc_num)
            info = accounts_master.get(key)

            if info is None:
                not_found.append(acc_num)
                results.append({
                    "Account Number":  acc_num,
                    "Account Name":    "",
                    "Opening Balance": None,
                    "Closing Balance": None,
                    "_mvmt":    None,
                    "_missing": True,
                })
                continue

            opening  = info["opening"]
            net_mvmt = movements_net.get(_norm(info["number"]), 0.0)
            closing  = opening + net_mvmt

            results.append({
                "Account Number":  info["number"],
                "Account Name":    info["name"],
                "Opening Balance": opening,
                "Closing Balance": closing,
                "_mvmt":    net_mvmt,
                "_missing": False,
            })

        # ── Step 5: debug summary ─────────────────────────────────────────
        st.divider()
        st.subheader("Debug Summary")

        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Accounts from 331",  len(pdf_accounts))
        d2.metric("B100 records",        f"{b100_count:,}")
        d3.metric("C100 records (total)", f"{c100_count:,}")
        d4.metric(f"Accounts with {TARGET_YEAR} movements", f"{len(movements_net):,}")

        if not_found:
            st.warning(
                f"{len(not_found)} account(s) not found in B100: "
                + ", ".join(not_found[:20])
            )

        # 5 matched examples
        sample = [r for r in results if not r["_missing"]][:5]
        if sample:
            with st.expander("Sample: first 5 matched accounts", expanded=True):
                for r in sample:
                    st.text(
                        f"  PDF={r['Account Number']:8s}  "
                        f"opening={str(r['Opening Balance']):>14}  "
                        f"movements={str(r['_mvmt']):>14}  "
                        f"closing={str(r['Closing Balance']):>14}"
                    )
        else:
            st.warning("No matched accounts — verify B100 account number position.")

        # ── Step 6: preview ───────────────────────────────────────────────
        import pandas as pd
        df = pd.DataFrame([{
            "Account Number":  r["Account Number"],
            "Account Name":    r["Account Name"],
            "Opening Balance": r["Opening Balance"],
            "Closing Balance": r["Closing Balance"],
        } for r in results])
        st.subheader("Results Preview")
        st.dataframe(df, use_container_width=True)

        # ── Step 7: Excel ─────────────────────────────────────────────────
        progress.progress(90, text="Building Excel…")

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Accounts"
        bold         = Font(bold=True)
        missing_fill = PatternFill("solid", fgColor="FFC7CE")

        ws.append(["Account Number", "Account Name", "Opening Balance", "Closing Balance"])
        for c in ws[1]: c.font = bold

        for r in results:
            ws.append([r["Account Number"], r["Account Name"],
                       r["Opening Balance"], r["Closing Balance"]])
            if r["_missing"]:
                for c in ws[ws.max_row]: c.fill = missing_fill

        matched = [r for r in results if not r["_missing"]]
        ws.append([])
        ws.append(["TOTAL", "",
                   sum(r["Opening Balance"] for r in matched),
                   sum(r["Closing Balance"] for r in matched)])
        for c in ws[ws.max_row]: c.font = bold

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 60
            )

        # Debug sheet
        ws_d = wb.create_sheet("Debug")
        ws_d.append(["Metric", "Value"])
        for c in ws_d[1]: c.font = bold
        ws_d.append(["Accounts from 331 PDF",         len(pdf_accounts)])
        ws_d.append(["B100 records in file",           b100_count])
        ws_d.append([f"C100 records (all)",            c100_count])
        ws_d.append([f"Accounts with {TARGET_YEAR} movements", len(movements_net)])
        ws_d.append(["Accounts not found in B100",     len(not_found)])
        ws_d.append([])
        ws_d.append(["Account Number", "Opening Balance", "Movement Sum", "Closing Balance"])
        for r in sample:
            ws_d.append([r["Account Number"], r["Opening Balance"],
                         r["_mvmt"], r["Closing Balance"]])
        for col in ws_d.columns:
            ws_d.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 50
            )

        # Save & open
        progress.progress(100, text="Done.")
        output_dir = _HERE / "output"
        output_dir.mkdir(exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = output_dir / f"Balances_{hcode}_{ts}.xlsx"
        wb.save(str(out_path))

        st.divider()
        st.success(f"Excel saved: `{out_path}`")
        try:
            os.startfile(str(out_path))
            st.info("Excel opened automatically.")
        except Exception as e:
            st.warning(f"Could not auto-open: {e}")

        with open(str(out_path), "rb") as fh:
            st.download_button(
                f"⬇ Download {out_path.name}", fh.read(),
                file_name=out_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    finally:
        try:
            os.unlink(pdf_path)
        except Exception:
            pass
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)
