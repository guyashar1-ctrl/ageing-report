"""
excel_generator.py - יצירת קובץ Excel מעוצב
=============================================
גיליון בודד RTL עם כותרות כחולות, זברה, שורת סיכום, freeze panes ו-auto-filter.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ageing_report.config.constants import EXCEL_HEADERS, SHEET_NAME


# --- צבעים ---
COLOR_HEADER_BG    = '1F4E79'   # כחול כהה - כותרות
COLOR_HEADER_FG    = 'FFFFFF'   # לבן - טקסט כותרות
COLOR_ROW_ODD      = 'DEEAF1'   # תכלת בהיר - שורות אי-זוגיות
COLOR_ROW_EVEN     = 'FFFFFF'   # לבן - שורות זוגיות
COLOR_CREDIT_FG    = 'C00000'   # אדום כהה - יתרת זכות (שלילי)
COLOR_DEBIT_FG     = '1F4E79'   # כחול כהה - יתרת חובה (חיובי)
COLOR_TOTAL_BG     = 'BDD7EE'   # תכלת בינוני - שורת סיכום
COLOR_TOTAL_FG     = '1F4E79'   # כחול כהה - טקסט סיכום


def _thin_border():
    s = Side(style='thin', color='B8CCE4')
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom_border():
    thin = Side(style='thin', color='B8CCE4')
    thick = Side(style='medium', color='1F4E79')
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def generate_excel(results):
    """
    יצירת קובץ Excel מתוצאות הגיול.

    Args:
        results: list[dict] מ-process_accounts()

    Returns:
        bytes: תוכן קובץ Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.rightToLeft = True

    border = _thin_border()

    # --- כותרות ---
    header_font = Font(bold=True, size=11, name='Arial', color=COLOR_HEADER_FG)
    header_fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thick_bottom_border()

    ws.row_dimensions[1].height = 32

    # --- שורות נתונים ---
    for row_idx, r in enumerate(results, 2):
        is_odd = (row_idx % 2 == 0)
        row_bg = COLOR_ROW_ODD if is_odd else COLOR_ROW_EVEN
        row_fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type='solid')

        closing = r['closing']
        closing_color = COLOR_CREDIT_FG if closing < 0 else COLOR_DEBIT_FG

        cells_data = [
            (1, r['acct_num'],           'center', None,         COLOR_DEBIT_FG),
            (2, r['name'],               'right',  None,         COLOR_DEBIT_FG),
            (3, closing,                 'center', '#,##0.00',   closing_color),
            (4, r['opening'],            'center', '#,##0.00',   COLOR_DEBIT_FG),
            (5, r['debt_start_date'],    'center', None,         COLOR_DEBIT_FG),
            (6, r['sum_formula'] or "",  'center', '#,##0.00',   COLOR_DEBIT_FG),
        ]
        for col, value, align, fmt, fg in cells_data:
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(size=10, name='Arial', color=fg)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.border = border
            if fmt:
                cell.number_format = fmt

        ws.row_dimensions[row_idx].height = 18

    # --- שורת סיכום ---
    total_row = len(results) + 2
    total_fill = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type='solid')
    total_font = Font(bold=True, size=10, name='Arial', color=COLOR_TOTAL_FG)
    total_border = _thick_bottom_border()

    total_closing = sum(r['closing'] for r in results)
    total_opening = sum(r['opening'] for r in results)

    totals_data = [
        (1, 'סה"כ',        'center', None,        COLOR_TOTAL_FG),
        (2, '',             'center', None,        COLOR_TOTAL_FG),
        (3, total_closing,  'center', '#,##0.00',  COLOR_CREDIT_FG if total_closing < 0 else COLOR_DEBIT_FG),
        (4, total_opening,  'center', '#,##0.00',  COLOR_CREDIT_FG if total_opening < 0 else COLOR_DEBIT_FG),
        (5, '',             'center', None,        COLOR_TOTAL_FG),
        (6, '',             'center', None,        COLOR_TOTAL_FG),
    ]
    for col, value, align, fmt, fg in totals_data:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True, size=10, name='Arial', color=fg)
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = total_border
        if fmt:
            cell.number_format = fmt

    ws.row_dimensions[total_row].height = 20

    # --- רוחב עמודות ---
    widths = {'A': 15, 'B': 40, 'C': 16, 'D': 16, 'E': 22, 'F': 50}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{len(results) + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
