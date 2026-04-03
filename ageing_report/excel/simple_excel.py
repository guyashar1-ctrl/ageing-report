"""
simple_excel.py
===============
Generates a simplified Excel workbook from SimpleAccountResult objects.

Workbook structure
------------------
Sheet 1 – "ריכוז"    (Main: Account Number, Name, Opening Balance, Closing Balance)
Sheet 2 – "debug"    (5 matched examples showing full extraction details)
Sheet 3 – "לוג"      (Warnings and errors)

Public API
----------
generate_simple_excel(
    results        : list[SimpleAccountResult],
    debug_samples  : list[str],          # raw B100 extraction debug lines
    c100_movements : list[C100Fixed],    # for debug examples
    b100_records   : dict,               # for debug examples
    header_code    : str,
    header_name    : str,
) -> bytes
"""

from __future__ import annotations

import io
from typing import Dict, List

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from parsers.bkmv_fixed_parser import B100Fixed, C100Fixed, get_c100_for_account
from processors.simple_balance import SimpleAccountResult
from utils.logger import get_accumulated_logs, get_logger

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
_WARN_FILL    = PatternFill("solid", fgColor="FFF2CC")
_ERROR_FILL   = PatternFill("solid", fgColor="FFE0E0")
_DEBUG_FILL   = PatternFill("solid", fgColor="E2EFDA")

_HEADER_FONT  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_BODY_FONT    = Font(name="Arial", size=10)
_BOLD_FONT    = Font(name="Arial", size=10, bold=True)
_WARN_FONT    = Font(name="Arial", size=10, color="C00000")

_RTL   = Alignment(horizontal="right", vertical="center", readingOrder=2)
_CTR   = Alignment(horizontal="center", vertical="center", readingOrder=2)
_LEFT  = Alignment(horizontal="left", vertical="center")

_THIN       = Side(style="thin", color="BFBFBF")
_THIN_BDR   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM_FMT    = '#,##0.00;[Red]-#,##0.00'


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_simple_excel(
    results:        List[SimpleAccountResult],
    debug_samples:  List[str],
    c100_movements: List[C100Fixed],
    b100_records:   Dict[str, B100Fixed],
    header_code:    str,
    header_name:    str,
) -> bytes:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    ws_main  = wb.create_sheet("ריכוז")
    ws_debug = wb.create_sheet("debug")
    ws_log   = wb.create_sheet("לוג")

    _build_main_sheet(ws_main, results, header_code, header_name)
    _build_debug_sheet(ws_debug, results, c100_movements, b100_records, debug_samples)
    _build_log_sheet(ws_log)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info("Simple Excel generated successfully")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1: ריכוז
# ---------------------------------------------------------------------------

def _build_main_sheet(ws, results, header_code, header_name):
    ws.sheet_view.rightToLeft = True

    # Title
    title = f"ריכוז גיול חשבונות – כותרת {header_code}  {header_name}"
    cell = ws.cell(1, 1, value=title)
    cell.font = Font(name="Arial", size=13, bold=True, color="1F4E79")
    cell.alignment = _RTL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Column headers
    headers = [
        "מספר חשבון",       # A
        "שם חשבון",         # B
        "יתרת פתיחה",       # C
        "סיכום תנועות",     # D
        "יתרת סגירה",       # E
    ]
    _write_header_row(ws, 2, headers)
    _set_col_widths(ws, [18, 40, 18, 18, 18])

    is_alt = False
    for i, r in enumerate(results, start=3):
        fill = _ALT_FILL if is_alt else None
        is_alt = not is_alt

        if not r.matched:
            fill = _WARN_FILL

        ws.cell(i, 1, r.pdf_account_number).alignment = _RTL
        ws.cell(i, 2, r.account_name).alignment        = _RTL

        for col, val in [(3, r.opening_balance), (4, r.movement_sum), (5, r.closing_balance)]:
            c = ws.cell(i, col, val)
            c.number_format = _NUM_FMT
            c.alignment = _RTL
            if not r.matched:
                c.font = _WARN_FONT

        if fill:
            for col in range(1, 6):
                ws.cell(i, col).fill = fill

        for col in range(1, 6):
            ws.cell(i, col).border = _THIN_BDR
            if ws.cell(i, col).font == Font():   # only set font if not already set
                ws.cell(i, col).font = _BODY_FONT

    # Totals row
    total_row = 3 + len(results)
    ws.cell(total_row, 2, "סה\"כ").font = _BOLD_FONT
    for col_letter, col_idx in [("C", 3), ("D", 4), ("E", 5)]:
        if total_row > 3:
            tc = ws.cell(total_row, col_idx)
            tc.value        = f"=SUM({col_letter}3:{col_letter}{total_row-1})"
            tc.number_format = _NUM_FMT
            tc.font          = _BOLD_FONT
            tc.alignment     = _RTL

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 2: debug
# ---------------------------------------------------------------------------

def _build_debug_sheet(
    ws,
    results:        List[SimpleAccountResult],
    c100_movements: List[C100Fixed],
    b100_records:   Dict[str, B100Fixed],
    debug_samples:  List[str],
):
    ws.sheet_view.rightToLeft = False   # debug sheet is LTR for readability

    row = 1

    # ── Section A: Matched account examples (up to 5) ───────────────────
    ws.cell(row, 1, "=== MATCHED ACCOUNT EXAMPLES (up to 5) ===").font = _BOLD_FONT
    ws.cell(row, 1).fill = _DEBUG_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    ex_headers = [
        "PDF Account#",
        "B100 Account#",
        "C100 Account# (sample)",
        "Account Name",
        "Opening Balance",
        "Movement Sum",
        "Closing Balance",
    ]
    _write_header_row(ws, row, ex_headers)
    _set_col_widths(ws, [18, 18, 24, 35, 18, 18, 18])
    row += 1

    matched_examples = [r for r in results if r.matched][:5]
    for r in matched_examples:
        # Find one sample C100 for this account
        c100s  = get_c100_for_account(c100_movements, r.pdf_account_number)
        c100_sample = ""
        if c100s:
            sample = c100s[0]
            c100_sample = (
                f"acct={sample.account_number}  "
                f"mvmt={sample.movement:+.2f}  "
                f"last_match={sample.debug_last_match!r}"
            )

        ws.cell(row, 1, r.pdf_account_number)
        ws.cell(row, 2, r.b100_account_number)
        ws.cell(row, 3, c100_sample)
        ws.cell(row, 4, r.account_name)
        ws.cell(row, 5, r.opening_balance).number_format = _NUM_FMT
        ws.cell(row, 6, r.movement_sum).number_format    = _NUM_FMT
        ws.cell(row, 7, r.closing_balance).number_format = _NUM_FMT

        for col in range(1, 8):
            ws.cell(row, col).border    = _THIN_BDR
            ws.cell(row, col).font      = _BODY_FONT
            ws.cell(row, col).alignment = _LEFT
            ws.cell(row, col).fill      = _ALT_FILL if row % 2 == 0 else None or PatternFill()

        row += 1

    if not matched_examples:
        ws.cell(row, 1, "No matched accounts found.").font = _WARN_FONT
        row += 1

    row += 1  # spacer

    # ── Section B: B100 field extraction details ─────────────────────────
    ws.cell(row, 1, "=== B100 EXTRACTION DEBUG (first 10 records) ===").font = _BOLD_FONT
    ws.cell(row, 1).fill = _DEBUG_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    ws.cell(row, 1, "Detail").font = _BOLD_FONT
    ws.column_dimensions["A"].width = 150
    row += 1

    for sample in debug_samples:
        ws.cell(row, 1, sample).font      = Font(name="Courier New", size=9)
        ws.cell(row, 1).alignment         = _LEFT
        ws.cell(row, 1).border            = _THIN_BDR
        row += 1

    if not debug_samples:
        ws.cell(row, 1, "No B100 records parsed.").font = _WARN_FONT
        row += 1

    row += 1  # spacer

    # ── Section C: C100 sample records ───────────────────────────────────
    ws.cell(row, 1, "=== C100 SAMPLE MOVEMENTS (first 10) ===").font = _BOLD_FONT
    ws.cell(row, 1).fill = _DEBUG_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    c100_headers = ["Account#", "Movement (signed)", "Last Match", "Amount Raw", "Raw Line (first 120)"]
    for col, h in enumerate(c100_headers, start=1):
        ws.cell(row, col, h).font   = _BOLD_FONT
        ws.cell(row, col).alignment = _LEFT
    _set_col_widths(ws, [14, 18, 22, 18, 120])
    row += 1

    for m in c100_movements[:10]:
        ws.cell(row, 1, m.account_number)
        ws.cell(row, 2, m.movement).number_format = _NUM_FMT
        ws.cell(row, 3, m.debug_last_match)
        ws.cell(row, 4, m.debug_amount_raw)
        ws.cell(row, 5, m.raw_line[:120].replace("\t", "→"))
        for col in range(1, 6):
            ws.cell(row, col).font      = Font(name="Courier New", size=9)
            ws.cell(row, col).alignment = _LEFT
            ws.cell(row, col).border    = _THIN_BDR
        row += 1


# ---------------------------------------------------------------------------
# Sheet 3: לוג
# ---------------------------------------------------------------------------

def _build_log_sheet(ws):
    ws.sheet_view.rightToLeft = True
    _write_header_row(ws, 1, ["רמה", "תאריך ושעה", "הודעה"])
    _set_col_widths(ws, [12, 20, 120])

    row = 2
    for level, message, timestamp in get_accumulated_logs():
        fill = _ERROR_FILL if level in ("ERROR", "CRITICAL") else (
            _WARN_FILL if level == "WARNING" else None
        )
        ws.cell(row, 1, level).alignment   = _CTR
        ws.cell(row, 2, timestamp).alignment = _CTR
        ws.cell(row, 3, message).alignment = _RTL
        if fill:
            for col in range(1, 4):
                ws.cell(row, col).fill = fill
        for col in range(1, 4):
            ws.cell(row, col).font   = _BODY_FONT
            ws.cell(row, col).border = _THIN_BDR
        row += 1

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, row: int, labels: list) -> None:
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row, col, label)
        c.fill      = _HEADER_FILL
        c.font      = _HEADER_FONT
        c.alignment = _RTL
        c.border    = _THIN_BDR
    ws.row_dimensions[row].height = 22


def _set_col_widths(ws, widths: list) -> None:
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
