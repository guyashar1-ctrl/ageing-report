"""
test_excel.py
=============
Standalone test: create a minimal Excel file and auto-open it.
Run directly:
    python test_excel.py

This is completely independent of the 331 / BKMVDATA logic.
It verifies:
  1. openpyxl can create a workbook
  2. The file can be written to disk
  3. os.startfile opens it in Excel
"""

import os
import sys
from datetime import date
from pathlib import Path

# ── 1. Setup output path ──────────────────────────────────────────────────
HERE       = Path(__file__).parent
OUTPUT_DIR = HERE / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT     = OUTPUT_DIR / "test_minimal.xlsx"

print(f"\n{'='*60}")
print("STAGE 1: openpyxl import")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    print(f"  ✓  openpyxl {openpyxl.__version__} imported OK")
except ImportError as e:
    print(f"  ✗  openpyxl not available: {e}")
    print("     Run:  python -m pip install openpyxl")
    sys.exit(1)

# ── 2. Build workbook ─────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("STAGE 2: build workbook")
try:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "גיול בדיקה"
    ws.sheet_view.rightToLeft = True

    # Header row
    headers = ["מספר חשבון", "שם חשבון", "יתרה נוכחית", "יתרת פתיחה",
               "סכום ח/ז", "תאריך תחילת חוב"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # 2 fake data rows
    rows = [
        ("10001", "לקוח בדיקה א",  5000.00,  3000.00, 5000.00, date(2025, 3, 15)),
        ("10002", "לקוח בדיקה ב", -2000.00, -1500.00, 2000.00, date(2025, 6, 1)),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="right")
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws.freeze_panes = "A2"
    print(f"  ✓  Workbook built: {ws.max_row - 1} data rows")
except Exception as e:
    print(f"  ✗  Workbook build failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# ── 3. Write to disk ──────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("STAGE 3: write to disk")
print(f"  Target path: {OUTPUT}")
try:
    wb.save(str(OUTPUT))
    if OUTPUT.exists():
        size = OUTPUT.stat().st_size
        print(f"  ✓  File created: {OUTPUT}  ({size:,} bytes)")
    else:
        print(f"  ✗  wb.save() did not raise but file does NOT exist at {OUTPUT}")
        sys.exit(1)
except Exception as e:
    print(f"  ✗  Write failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# ── 4. Auto-open ──────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("STAGE 4: auto-open with OS default app")
try:
    if sys.platform == "win32":
        os.startfile(str(OUTPUT))
        print(f"  ✓  os.startfile() called – Excel should open shortly")
    else:
        import subprocess
        subprocess.Popen(["open" if sys.platform == "darwin" else "xdg-open", str(OUTPUT)])
        print(f"  ✓  open/xdg-open called")
except Exception as e:
    print(f"  ✗  Auto-open failed: {e}")
    print(f"     File is at: {OUTPUT}")
    print(f"     You can open it manually.")

print(f"\n{'='*60}")
print("ALL STAGES COMPLETE")
print(f"File path: {OUTPUT}")
print(f"{'='*60}\n")
