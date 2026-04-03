#!/usr/bin/env python3
"""
דוח גיול לקוחות - יצירה אוטומטית מדוח 331 ומבנה אחיד
======================================================
קלט:
  1. קובץ PDF של דוח 331 (פירוטים)
  2. קובץ BKMVDATA.TXT מתוך מבנה אחיד (ZIP)

פלט:
  קובץ Excel עם: מספר חשבון, שם חשבון, יתרה נוכחית, יתרת פתיחה, תאריך תחילת חוב, פירוט סכימה
"""

import re
import sys
import os
from collections import defaultdict

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("חסרות חבילות. יש להריץ:")
    print("  pip3 install pdfplumber openpyxl")
    sys.exit(1)


# ===================== הגדרות =====================
PDF_PATH = "ספרינג אווירה וריח 12-2025_דוח 331.pdf"
BKMV_PATH = "bkmv_data/BKMVDATA.TXT"
OUTPUT_PATH = "דוח_גיול_לקוחות_ספרינג_2025.xlsx"
SECTION_CODE = "1342"

# מחרוזות זיהוי לסעיף 1342 בטקסט הויזואלי של ה-PDF
SECTION_START = "ץראב םיחותפ תובוח 1342"  # "1342 חובות פתוחים בארץ" בסדר ויזואלי
SECTION_CONTINUE = "(ךשמה)"                 # "(המשך)" בסדר ויזואלי


def parse_pdf_section_1342(pdf_path):
    """חילוץ רשימת חשבונות מסעיף 1342 בדוח 331"""
    print(f"[1/5] קורא PDF: {pdf_path}")
    pdf_accounts = {}

    with pdfplumber.open(pdf_path) as pdf:
        in_section = False
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                # תחילת סעיף
                if SECTION_START in line and 'כ"הס' not in line:
                    in_section = True
                    continue
                # שורות המשך
                if in_section and SECTION_CONTINUE in line:
                    continue
                # סוף סעיף (שורת סה"כ)
                if in_section and SECTION_CODE in line and 'כ"הס' in line:
                    in_section = False
                    continue
                if in_section:
                    m = re.match(
                        r'^([\d,]+\.\d{2})\s+(ח|ז)\s+(\d+)\s+(.+?)\s+(\d{3,6})$',
                        line.strip()
                    )
                    if m:
                        acct_num = int(m.group(5))
                        pdf_accounts[acct_num] = {
                            'balance_pdf': float(m.group(1).replace(',', '')),
                            'bal_type': m.group(2),
                            'name_visual': m.group(4).strip()
                        }

    print(f"      נמצאו {len(pdf_accounts)} חשבונות בסעיף {SECTION_CODE}")
    return pdf_accounts


def parse_b11_records(bkmv_path, target_accounts):
    """קריאת רשומות B11 - כרטיסיות חשבון עם יתרות"""
    print(f"[2/5] קורא כרטיסיות חשבון (B11)...")
    b11_data = {}

    with open(bkmv_path, 'rb') as f:
        for line in f:
            if not line.startswith(b'B11'):
                continue
            d = line.rstrip(b'\r\n\x85').decode('cp862', errors='replace')

            acct_str = d[22:37].strip()
            try:
                acct_num = int(acct_str)
            except ValueError:
                continue

            if acct_num not in target_accounts:
                continue

            name_visual = d[37:87].strip()
            name_logical = name_visual[::-1]  # היפוך לסדר לוגי

            # חילוץ סכומים: סימן + 14 ספרות, באגורות
            amts = re.findall(r'[+-]\d{14}', d[270:])
            opening = int(amts[0]) / 100 if len(amts) >= 1 else 0
            total_debits = int(amts[1]) / 100 if len(amts) >= 2 else 0
            total_credits = int(amts[2]) / 100 if len(amts) >= 3 else 0

            b11_data[acct_num] = {
                'name': name_logical,
                'opening_balance': opening,
                'total_debits': total_debits,
                'total_credits': total_credits,
            }

    print(f"      התאמה: {len(b11_data)}/{len(target_accounts)} חשבונות")
    return b11_data


def parse_b1_transactions(bkmv_path, target_accounts):
    """קריאת רשומות B1 - תנועות יומן"""
    print(f"[3/5] קורא תנועות יומן (B1)...")
    transactions = defaultdict(list)

    with open(bkmv_path, 'rb') as f:
        for line in f:
            if not (line.startswith(b'B1') and not line.startswith(b'B11')):
                continue
            d = line.rstrip(b'\r\n\x85').decode('cp862', errors='replace')

            acct_str = d[172:187].strip()
            try:
                acct_num = int(acct_str)
            except ValueError:
                continue

            if acct_num not in target_accounts:
                continue

            date_str = d[156:164]           # YYYYMMDD
            dc = d[202:203].strip()         # 1=חובה, 2=זכות

            amt_match = re.search(r'[+-]\d{14}', d[203:])
            if not amt_match:
                continue
            amount = int(amt_match.group()) / 100

            if dc in ('1', '2') and amount != 0:
                transactions[acct_num].append((date_str, dc, amount))

    print(f"      נטענו תנועות עבור {len(transactions)} חשבונות")
    return transactions


def calculate_aging(closing, acct_txns, opening):
    """
    חישוב תאריך תחילת חוב:
    סוכמים תנועות מסוף השנה לתחילתה (רק חובה/רק זכות בהתאם ליתרה)
    עד שהסכום המצטבר עובר את יתרת הסגירה.
    """
    debt_start_date = ""
    sum_parts = []

    if closing > 0.005:
        # יתרת חובה: סוכמים תנועות חובה (dc=1, סכום חיובי) מהסוף להתחלה
        relevant = [(d, amt) for d, dc, amt in acct_txns if dc == '1' and amt > 0]
        relevant.sort(key=lambda x: x[0], reverse=True)
        target = closing
    elif closing < -0.005:
        # יתרת זכות: סוכמים תנועות זכות (dc=2, סכום חיובי) מהסוף להתחלה
        relevant = [(d, amt) for d, dc, amt in acct_txns if dc == '2' and amt > 0]
        relevant.sort(key=lambda x: x[0], reverse=True)
        target = abs(closing)
    else:
        return "", []

    cumulative = 0
    for date_str, amt in relevant:
        cumulative += amt
        sum_parts.append(amt)
        if cumulative >= target - 0.005:
            debt_start_date = f"{date_str[6:8]}/{date_str[4:6]}/{date_str[0:4]}"
            break

    if not debt_start_date and sum_parts:
        if (closing > 0 and opening > 0) or (closing < 0 and opening < 0):
            debt_start_date = "כולל יתרת פתיחה"
            sum_parts.append(abs(opening))
        else:
            debt_start_date = "מורכב ממספר יתרות"

    return debt_start_date, sum_parts


def generate_excel(results, output_path):
    """יצירת קובץ Excel עם דוח הגיול"""
    print(f"[5/5] יוצר Excel: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "גיול לקוחות"
    ws.sheet_view.rightToLeft = True

    # כותרות
    headers = ['מספר חשבון', 'שם חשבון', 'יתרה נוכחית', 'יתרת פתיחה', 'תאריך תחילת חוב', 'פירוט סכימה']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=12, name='Arial', color='FFFFFF')
    data_font = Font(size=11, name='Arial')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # שורות נתונים
    for row_idx, r in enumerate(results, 2):
        cells = [
            (1, r['acct_num'], 'center', None),
            (2, r['name'], 'right', None),
            (3, r['closing'], 'center', '#,##0.00'),
            (4, r['opening'], 'center', '#,##0.00'),
            (5, r['debt_start_date'], 'center', None),
            (6, r['sum_formula'] or "", 'center', '#,##0.00'),
        ]
        for col, value, align, fmt in cells:
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal=align)
            cell.border = thin_border
            if fmt:
                cell.number_format = fmt

    # רוחב עמודות
    widths = {'A': 15, 'B': 45, 'C': 18, 'D': 18, 'E': 22, 'F': 50}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{len(results) + 1}"

    wb.save(output_path)
    return len(results)


def main():
    # שלב 1: חילוץ חשבונות מהדוח 331
    pdf_accounts = parse_pdf_section_1342(PDF_PATH)
    if not pdf_accounts:
        print("שגיאה: לא נמצאו חשבונות בסעיף 1342")
        sys.exit(1)
    target_accounts = set(pdf_accounts.keys())

    # שלב 2: קריאת B11 - כרטיסיות
    b11_data = parse_b11_records(BKMV_PATH, target_accounts)

    # שלב 3: קריאת B1 - תנועות
    transactions = parse_b1_transactions(BKMV_PATH, target_accounts)

    # שלב 4: חישוב גיול
    print(f"[4/5] מחשב גיול...")
    results = []

    for acct_num in sorted(pdf_accounts.keys()):
        pdf_info = pdf_accounts[acct_num]
        b11 = b11_data.get(acct_num, {})

        name = b11.get('name', pdf_info['name_visual'][::-1])
        opening = b11.get('opening_balance', 0)
        total_debits = b11.get('total_debits', 0)
        total_credits = b11.get('total_credits', 0)
        closing = opening + total_debits - total_credits

        acct_txns = transactions.get(acct_num, [])
        debt_start_date, sum_parts = calculate_aging(closing, acct_txns, opening)

        # בניית נוסחת SUM לאקסל
        sum_formula = ""
        if sum_parts:
            sum_formula = "=" + "+".join(f"{x:.2f}" for x in sum_parts)

        results.append({
            'acct_num': acct_num,
            'name': name,
            'closing': closing,
            'opening': opening,
            'debt_start_date': debt_start_date,
            'sum_formula': sum_formula,
        })

    # שלב 5: יצירת Excel
    row_count = generate_excel(results, OUTPUT_PATH)

    # סיכום
    debit_count = sum(1 for r in results if r['closing'] > 0.005)
    credit_count = sum(1 for r in results if r['closing'] < -0.005)
    with_date = sum(1 for r in results if re.match(r'\d{2}/\d{2}/\d{4}', r['debt_start_date']))
    total_closing = sum(r['closing'] for r in results)

    print(f"\n{'='*50}")
    print(f"  הושלם בהצלחה!")
    print(f"  קובץ: {OUTPUT_PATH}")
    print(f"  חשבונות: {row_count}")
    print(f"  יתרות חובה: {debit_count}")
    print(f"  יתרות זכות: {credit_count}")
    print(f"  עם תאריך מדויק: {with_date}")
    print(f"  סה\"כ יתרות: {total_closing:,.2f}")
    print(f"{'='*50}")


if __name__ == '__main__':
    main()
