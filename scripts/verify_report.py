#!/usr/bin/env python3
"""
סקריפט אימות: השוואת דוח הגיול לדוח 331
=========================================
בודק שהאקסל שנוצר תואם את הנתונים בדוח 331:
- ספירת חשבונות
- סה"כ יתרות
- דגימה של חשבונות ספציפיים

שימוש:
    python3 scripts/verify_report.py <PDF_PATH> <EXCEL_PATH>
"""

import sys
import re

try:
    import pdfplumber
    from openpyxl import load_workbook
except ImportError:
    print("חסרות חבילות. יש להריץ:")
    print("  pip3 install pdfplumber openpyxl")
    sys.exit(1)


def extract_pdf_1342(pdf_path):
    """חילוץ חשבונות מסעיף 1342"""
    accounts = {}
    with pdfplumber.open(pdf_path) as pdf:
        in_section = False
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                if 'ץראב םיחותפ תובוח 1342' in line and 'כ"הס' not in line:
                    in_section = True
                    continue
                if in_section and '(ךשמה)' in line:
                    continue
                if in_section and '1342' in line and 'כ"הס' in line:
                    # Extract total
                    m = re.match(r'^([\d,]+\.\d{2})', line.strip())
                    if m:
                        accounts['_total'] = float(m.group(1).replace(',', ''))
                    in_section = False
                    continue
                if in_section:
                    m = re.match(
                        r'^([\d,]+\.\d{2})\s+(ח|ז)\s+(\d+)\s+(.+?)\s+(\d{3,6})$',
                        line.strip()
                    )
                    if m:
                        acct_num = int(m.group(5))
                        balance = float(m.group(1).replace(',', ''))
                        bal_type = m.group(2)  # ח=חובה, ז=זכות
                        sign = 1 if bal_type == 'ח' else -1
                        accounts[acct_num] = balance * sign
    return accounts


def load_excel(excel_path):
    """טעינת נתונים מהאקסל"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    accounts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            acct_num = int(row[0])
            closing = float(row[2]) if row[2] else 0
            accounts[acct_num] = closing
    return accounts


def verify(pdf_path, excel_path):
    print("=" * 60)
    print("אימות דוח גיול מול דוח 331")
    print("=" * 60)

    # טעינה
    print(f"\nקורא PDF: {pdf_path}")
    pdf_data = extract_pdf_1342(pdf_path)
    pdf_total = pdf_data.pop('_total', 0)
    print(f"  חשבונות ב-PDF: {len(pdf_data)}")
    print(f"  סה\"כ ב-PDF: {pdf_total:,.2f}")

    print(f"\nקורא Excel: {excel_path}")
    excel_data = load_excel(excel_path)
    print(f"  חשבונות באקסל: {len(excel_data)}")
    excel_total = sum(excel_data.values())
    print(f"  סה\"כ באקסל: {excel_total:,.2f}")

    # בדיקות
    print(f"\n{'=' * 60}")
    print("תוצאות אימות:")
    print("-" * 60)

    errors = 0

    # 1. ספירת חשבונות
    if len(pdf_data) == len(excel_data):
        print(f"  ✓ ספירת חשבונות: {len(pdf_data)} = {len(excel_data)}")
    else:
        print(f"  ✗ ספירת חשבונות: PDF={len(pdf_data)}, Excel={len(excel_data)}")
        errors += 1
        # מצא חשבונות חסרים
        pdf_only = set(pdf_data.keys()) - set(excel_data.keys())
        excel_only = set(excel_data.keys()) - set(pdf_data.keys())
        if pdf_only:
            print(f"    ב-PDF בלבד: {pdf_only}")
        if excel_only:
            print(f"    באקסל בלבד: {excel_only}")

    # 2. סה"כ יתרות
    if abs(excel_total - pdf_total) < 0.1:
        print(f"  ✓ סה\"כ יתרות: {excel_total:,.2f} ≈ {pdf_total:,.2f}")
    else:
        print(f"  ✗ סה\"כ יתרות: Excel={excel_total:,.2f}, PDF={pdf_total:,.2f}")
        print(f"    הפרש: {excel_total - pdf_total:,.2f}")
        errors += 1

    # 3. דגימת חשבונות
    print(f"\n  דגימת חשבונות:")
    mismatches = 0
    for acct_num in sorted(pdf_data.keys()):
        pdf_bal = pdf_data[acct_num]
        excel_bal = excel_data.get(acct_num)
        if excel_bal is None:
            continue
        if abs(pdf_bal - excel_bal) > 0.01:
            mismatches += 1
            if mismatches <= 10:
                print(f"    ✗ חשבון {acct_num}: PDF={pdf_bal:,.2f}, Excel={excel_bal:,.2f}")

    if mismatches == 0:
        print(f"    ✓ כל {len(pdf_data)} החשבונות תואמים")
    else:
        print(f"    ✗ {mismatches} חשבונות עם אי-התאמה")
        errors += 1

    # סיכום
    print(f"\n{'=' * 60}")
    if errors == 0:
        print("  ✓ כל הבדיקות עברו בהצלחה!")
    else:
        print(f"  ✗ נמצאו {errors} בעיות")
    print("=" * 60)

    return errors == 0


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("שימוש: python3 verify_report.py <PDF_PATH> <EXCEL_PATH>")
        sys.exit(1)

    success = verify(sys.argv[1], sys.argv[2])
    sys.exit(0 if success else 1)
